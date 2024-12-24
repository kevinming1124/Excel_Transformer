from PyQt5 import QtWidgets, QtCore, QtGui

import mainWindow
from openpyxl import load_workbook, Workbook

class main(QtWidgets.QWidget, mainWindow.Ui_ExcelTransformer):
    def __init__(self):
        super(main, self).__init__()
        self.ui = mainWindow.Ui_ExcelTransformer()
        self.ui.setupUi(self)
        self.filePath = ""
        self.checkboxes = []
        self.ui.openFile.clicked.connect(self.openFile)
        self.ui.saveFile.clicked.connect(self.saveFile)

    def openFile(self):
        filePath , _ = QtWidgets.QFileDialog.getOpenFileNames()
        if not filePath:
            return
        self.filePath = filePath[0]
        self.wb = load_workbook(self.filePath)
        self.ws = self.wb.active
        cell = self.ws.cell(row = 4, column = 1)
        list_of_values = cell.value.split(';')
        self.clear_layout(self.ui.ckBoxes)
        for item in list_of_values:
            checkbox = QtWidgets.QCheckBox(f"{item}", self)
            self.ui.ckBoxes.addWidget(checkbox)
            self.checkboxes.append(checkbox)
        self.setLayout(self.ui.ckBoxes)

    def saveFile(self):
        for row in self.ws.iter_rows(min_row=4, max_row=self.ws.max_row, min_col=1, max_col=self.ws.max_column):
            for cell in row:
                # split the value of the cell with a ; and store it in a list
                list_of_values = cell.value.split(';')
                # clear the cell
                cell.value = None
                # write the values in the list to the cell
                for value in list_of_values:
                    cell.value = value
                    cell = cell.offset(row=0, column=1)
        selected_items = []
        for checkbox in self.checkboxes:
            if checkbox.isChecked():
                selected_items.append(checkbox.text())
        header = [cell.value for cell in next(self.ws.iter_rows(min_row=4, max_row=4))]
        column_indices = {name: index + 1 for index, name in enumerate(header)}
        keep_columns = [column_indices[name] for name in selected_items if name in column_indices]
        new_workbook = Workbook()
        new_sheet = new_workbook.active
        new_sheet.title = "Filtered Data"
        for row in self.ws.iter_rows():
            if row[0].row < 4:
                new_row = [cell.value for cell in row]
                new_sheet.append(new_row)
                continue
            new_row = [row[col - 1].value for col in keep_columns]
            new_sheet.append(new_row)
        options = QtWidgets.QFileDialog.Options()
        file_path, _ = QtWidgets.QFileDialog.getSaveFileName(
            self,
            "Save File",                
            "",                         
            "Excel Files (*.xlsx);;All Files (*)",
            options=options
        )
        new_workbook.save(file_path)
        
    def updateckBoxes(self):
        self.ui.ckBoxes.addWidget(QtWidgets.QCheckBox("test"))
        
    def clear_layout(self, layout):
        while layout.count():
            item = layout.takeAt(0)  # 取出佈局中的第一個項目
            widget = item.widget()  # 檢查是否是 widget
            if widget is not None:
                widget.deleteLater()
                
if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    window = main()
    window.show()
    sys.exit(app.exec_())